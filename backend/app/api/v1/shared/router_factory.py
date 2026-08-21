import csv
import io
import inspect
import re
from typing import Any, Callable, Generic, List, Optional, Type, TypeVar

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile, status
from pydantic import BaseModel, ValidationError
from sqlalchemy import func, select
from sqlalchemy.orm import class_mapper

from app.schemas.shared.base import (
    APIResponse,
    BulkOperationRequest,
    FilterRequest,
    PaginationRequest,
    PaginationResponse,
    SearchRequest,
    SearchResponse,
)
from app.repositories.interfaces.base import BaseRepository, RepositoryError

from app.api.v1.shared.utils import _is_relationship_attribute, apply_patch, filter_items

TModel = TypeVar("TModel")
TCreate = TypeVar("TCreate", bound=BaseModel)
TUpdate = TypeVar("TUpdate", bound=BaseModel)
TDetail = TypeVar("TDetail", bound=BaseModel)
TList = TypeVar("TList", bound=BaseModel)


def build_crud_router(
    prefix: str,
    tags: list[str],
    repository_dependency: Callable[..., BaseRepository[Any]],
    service_dependency: Callable[..., Any],
    model_class: type[TModel],
    create_schema: type[TCreate],
    update_schema: type[TUpdate],
    detail_schema: type[TDetail],
    list_schema: type[TList],
    bulk_update_schema: type[BaseModel],
    extra_routes: Optional[Callable[[APIRouter], None]] = None,
    entity_name: Optional[str] = None,
) -> APIRouter:
    router = APIRouter(prefix=prefix, tags=tags)
    entity_label = entity_name or tags[0].rstrip("s")
    if not entity_label:
        entity_label = tags[0]

    def _orm_to_dict(obj):
        try:
            cols = [c.key for c in class_mapper(obj.__class__).columns]
            return {k: getattr(obj, k) for k in cols}
        except Exception:
            return {k: v for k, v in vars(obj).items() if not k.startswith("_")}

    def _serialize_related_value(value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (str, int, float, bool, dict, list, tuple)):
            return value
        # Prefer common display attributes from related models.
        for attr in ('name', 'title', 'code', 'employee_code', 'designation', 'department', 'id'):
            if hasattr(value, attr):
                try:
                    candidate = getattr(value, attr)
                except Exception:
                    continue
                if candidate is not None:
                    return candidate
        try:
            return str(value)
        except Exception:
            return None

    def _normalize_enum_value(key: str, value: Any) -> Any:
        if not isinstance(value, str):
            return value
        try:
            column = class_mapper(model_class).columns[key]
        except Exception:
            return value
        col_type = getattr(column, 'type', None)
        allowed_values = getattr(col_type, 'enums', None)
        if not allowed_values:
            return value
        for allowed in allowed_values:
            if isinstance(allowed, str) and allowed.lower() == value.lower():
                return allowed
        return value

    def _to_snake_case(value: str) -> str:
        if not isinstance(value, str):
            return value
        value = value.strip()
        value = value.replace(' ', '_').replace('-', '_')
        value = re.sub(r'(.)([A-Z][a-z]+)', r'\1_\2', value)
        value = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', value)
        return re.sub(r'__+', '_', value).lower()

    async def _read_import_rows(file: UploadFile) -> list[dict[str, Any]]:
        raw_data = await file.read()
        filename = (file.filename or '').lower()

        if filename.endswith('.csv'):
            text = raw_data.decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(text), skipinitialspace=True)
            rows = []
            for row in reader:
                cleaned = {
                    _to_snake_case(str(key)): (value.strip() if isinstance(value, str) else value)
                    for key, value in row.items()
                    if key and str(key).strip()
                }
                if any(value not in (None, '') for value in cleaned.values()):
                    rows.append(cleaned)
            return rows

        if filename.endswith('.xlsx'):
            try:
                import openpyxl
            except ImportError as exc:
                raise HTTPException(
                    status_code=status.HTTP_501_NOT_IMPLEMENTED,
                    detail='Excel import requires openpyxl on the backend.',
                ) from exc
            workbook = openpyxl.load_workbook(io.BytesIO(raw_data), data_only=True)
            sheet = workbook.active
            rows = []
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return rows
            headers = [str(value).strip() if value is not None else '' for value in header_row]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                cleaned = {}
                for header, cell_value in zip(headers, row):
                    if not header or not str(header).strip():
                        continue
                    value = cell_value
                    if isinstance(value, str):
                        value = value.strip()
                    cleaned[_to_snake_case(str(header))] = value
                if any(value not in (None, '') for value in cleaned.values()):
                    rows.append(cleaned)
            return rows

        if filename.endswith('.xls'):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail='Legacy .xls format is not supported. Please save the file as .xlsx or .csv.',
            )

        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail='Unsupported file type. Please upload a CSV or XLSX file.',
        )

    def _orm_to_dict_for_schema(obj, schema_type: type[BaseModel]):
        data = _orm_to_dict(obj)
        # Fill missing schema fields from object attributes (synonyms/properties)
        try:
            schema_fields = getattr(schema_type, 'model_fields', {})
            for field in schema_fields:
                if field not in data:
                    # try attribute on object
                    if hasattr(obj, field):
                        try:
                            value = getattr(obj, field)
                            data[field] = _serialize_related_value(value)
                            continue
                        except Exception:
                            pass
                    # common synonyms: map title -> name
                    if field == 'name' and 'title' in data:
                        data['name'] = data['title']
            # Convert relationship objects to serializable forms.
            for key, value in list(data.items()):
                if not isinstance(value, (str, int, float, bool, dict, list, tuple)):
                    data[key] = _serialize_related_value(value)
        except Exception:
            pass
        return data

    @router.get(
        "",
        response_model=APIResponse[PaginationResponse[list_schema]],
        summary=f"List {tags[0]}",
        description=f"Retrieve a paginated list of {tags[0]} records.",
    )
    async def list_entities(
        pagination: PaginationRequest = Depends(),
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        items, total = await repository.paginate(pagination.page, pagination.page_size)
        response = PaginationResponse[list_schema](
            items=[list_schema.model_validate(_orm_to_dict_for_schema(item, list_schema)) for item in items],
            total=total,
            page=pagination.page,
            page_size=pagination.page_size,
            pages=(total + pagination.page_size - 1) // pagination.page_size,
        )
        return APIResponse(data=response)

    @router.post(
        "",
        response_model=APIResponse[detail_schema],
        status_code=status.HTTP_201_CREATED,
        summary=f"Create {entity_label}",
        description=f"Create a new {entity_label} record.",
    )
    async def create_entity(
        payload: Any,
        request: Request,
        repository: BaseRepository[Any] = Depends(repository_dependency),
        service: Any = Depends(service_dependency),
    ):
        payload_data = payload.model_dump()

        # Helper: resolve a relationship display value to a foreign-key id
        def _resolve_relationship(key: str, value: Any, repo_session) -> Optional[int]:
            try:
                prop = class_mapper(model_class).get_property(key)
            except Exception:
                return None
            # related mapped class
            related_class = prop.mapper.class_
            # find local fk column name
            local_cols = list(prop.local_columns)
            if not local_cols:
                return None
            fk_name = local_cols[0].key

            # if value already looks like an id, return it
            if value is None:
                return None
            if isinstance(value, int):
                return int(value)
            if isinstance(value, str) and value.isdigit():
                return int(value)

            if repo_session is None:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Cannot resolve relation '{key}' without DB session")

            # try common lookup fields
            for candidate in ("code", "title", "name", "teacher_code", "employee_code", "roll_number", "admission_no"):
                if hasattr(related_class, candidate):
                    field = getattr(related_class, candidate)
                    if isinstance(value, str):
                        stmt = select(related_class).where(func.lower(field) == value.lower())
                    else:
                        stmt = select(related_class).where(field == value)
                    result = repo_session.execute(stmt).scalar_one_or_none()
                    if result is not None:
                        return int(getattr(result, "id"))

            # Attempt to create a new related lookup record when not found (helps free-text lookups from UI)
            try:
                for candidate in ("title", "name", "code"):
                    if hasattr(related_class, candidate):
                        # create minimal instance using the candidate attribute
                        new_obj = related_class(**{candidate: value})
                        repo_session.add(new_obj)
                        repo_session.flush()
                        return int(getattr(new_obj, "id"))
            except Exception:
                # creation failed — fall through to error
                pass

            # not resolved
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Related {related_class.__name__} not found for {key}: {value}")

        async def _call_service_method(method_name: str):
            method = getattr(service, method_name)
            method_signature = inspect.signature(method)
            if any(param.kind == inspect.Parameter.VAR_KEYWORD for param in method_signature.parameters.values()):
                kwargs = dict(payload_data)
                if 'request' in method_signature.parameters:
                    kwargs['request'] = request
                return await method(**kwargs)
            filtered_payload = {
                key: value
                for key, value in payload_data.items()
                if key in method_signature.parameters
            }
            if 'request' in method_signature.parameters:
                filtered_payload['request'] = request
            return await method(**filtered_payload)

        service_result = None
        if hasattr(service, "enroll_student"):
            service_result = await _call_service_method("enroll_student")
        elif hasattr(service, "create_application"):
            service_result = await _call_service_method("create_application")
        elif hasattr(service, "join_employee"):
            service_result = await _call_service_method("join_employee")
        elif hasattr(service, "create_invoice"):
            service_result = await _call_service_method("create_invoice")
        elif hasattr(service, "post_entry"):
            service_result = await _call_service_method("post_entry")
        elif hasattr(service, "allocate_room"):
            service_result = await _call_service_method("allocate_room")
        elif hasattr(service, "validate_stock"):
            service_result = await _call_service_method("validate_stock")
        elif hasattr(service, "issue_book") and all(
            parameter in payload_data
            for parameter in inspect.signature(service.issue_book).parameters
            if inspect.signature(service.issue_book).parameters[parameter].default is inspect.Parameter.empty
        ):
            service_result = await _call_service_method("issue_book")
        elif hasattr(service, "dispatch"):
            service_result = await _call_service_method("dispatch")
        elif hasattr(service, "create_request"):
            service_result = await _call_service_method("create_request")
        elif hasattr(service, "allocate_vehicle") and all(
            parameter in payload_data
            for parameter in inspect.signature(service.allocate_vehicle).parameters
            if inspect.signature(service.allocate_vehicle).parameters[parameter].default is inspect.Parameter.empty
        ):
            service_result = await _call_service_method("allocate_vehicle")
        elif hasattr(service, "create"):
            service_result = await _call_service_method("create")

        if isinstance(service_result, dict):
            payload_data = service_result
        elif isinstance(service_result, model_class):
            created = await repository.create(service_result)
            return APIResponse(data=detail_schema.model_validate(_orm_to_dict_for_schema(created, detail_schema)), message="Created")

        # Build payload mapping: convert relationship display values to fk columns
        filtered_payload = {}
        repo_session = getattr(repository, "session", None)
        for key, value in payload_data.items():
            if _is_relationship_attribute(model_class, key):
                fk_id = _resolve_relationship(key, value, repo_session)
                if fk_id is not None:
                    # map to fk column name
                    try:
                        prop = class_mapper(model_class).get_property(key)
                        local_cols = list(prop.local_columns)
                        if local_cols:
                            fk_name = local_cols[0].key
                            filtered_payload[fk_name] = fk_id
                    except Exception:
                        # fallback: skip relationship attribute
                        continue
                continue
            filtered_payload[key] = _normalize_enum_value(key, value)
        # Map common schema->model synonyms (e.g., schema uses `name` while model uses `title`)
        if 'name' in filtered_payload and not hasattr(model_class, 'name') and hasattr(model_class, 'title'):
            filtered_payload['title'] = filtered_payload.pop('name')

        entity = model_class(**filtered_payload)
        created = await repository.create(entity)
        return APIResponse(data=detail_schema.model_validate(_orm_to_dict_for_schema(created, detail_schema)), message="Created")

    create_entity.__annotations__["payload"] = create_schema

    if extra_routes is not None:
        extra_routes(router)

    @router.get(
        "/{entity_id}",
        response_model=APIResponse[detail_schema],
        summary=f"Get {entity_label}",
        description=f"Retrieve a single {entity_label} by ID.",
    )
    async def get_entity(
        entity_id: int,
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        entity = await repository.get_by_id(entity_id)
        if entity is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Resource not found")
        return APIResponse(data=detail_schema.model_validate(_orm_to_dict_for_schema(entity, detail_schema)))

    @router.put(
        "/{entity_id}",
        response_model=APIResponse[detail_schema],
        summary=f"Replace {entity_label}",
        description=f"Replace a {entity_label} record by ID.",
    )
    async def replace_entity(
        entity_id: int,
        payload: Any,
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        existing = await repository.get_by_id(entity_id)
        if existing is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Resource not found")
        update_data = payload.model_dump(exclude_unset=True)
        # resolve relationships in update_data
        repo_session = getattr(repository, "session", None)
        resolved = {}
        for key, value in update_data.items():
            if _is_relationship_attribute(model_class, key):
                fk_id = _resolve_relationship(key, value, repo_session)
                if fk_id is not None:
                    try:
                        prop = class_mapper(model_class).get_property(key)
                        local_cols = list(prop.local_columns)
                        if local_cols:
                            fk_name = local_cols[0].key
                            resolved[fk_name] = fk_id
                    except Exception:
                        continue
                continue
            resolved[key] = _normalize_enum_value(key, value)
        # map common schema->model synonyms for updates as well
        if 'name' in resolved and not hasattr(model_class, 'name') and hasattr(model_class, 'title'):
            resolved['title'] = resolved.pop('name')
        apply_patch(existing, resolved)
        await repository._commit()
        await repository._refresh(existing)
        return APIResponse(data=detail_schema.model_validate(_orm_to_dict_for_schema(existing, detail_schema)), message="Updated")

    replace_entity.__annotations__["payload"] = update_schema

    @router.patch(
        "/{entity_id}",
        response_model=APIResponse[detail_schema],
        summary=f"Update {entity_label}",
        description=f"Update one or more fields for a {entity_label}.",
    )
    async def patch_entity(
        entity_id: int,
        payload: Any,
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        existing = await repository.get_by_id(entity_id)
        if existing is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Resource not found")
        update_data = payload.model_dump(exclude_unset=True)
        repo_session = getattr(repository, "session", None)
        resolved = {}
        for key, value in update_data.items():
            if _is_relationship_attribute(model_class, key):
                fk_id = _resolve_relationship(key, value, repo_session)
                if fk_id is not None:
                    try:
                        prop = class_mapper(model_class).get_property(key)
                        local_cols = list(prop.local_columns)
                        if local_cols:
                            fk_name = local_cols[0].key
                            resolved[fk_name] = fk_id
                    except Exception:
                        continue
                continue
            resolved[key] = _normalize_enum_value(key, value)
        # map common schema->model synonyms for patch updates as well
        if 'name' in resolved and not hasattr(model_class, 'name') and hasattr(model_class, 'title'):
            resolved['title'] = resolved.pop('name')
        apply_patch(existing, resolved)
        await repository._commit()
        await repository._refresh(existing)
        return APIResponse(data=detail_schema.model_validate(_orm_to_dict_for_schema(existing, detail_schema)), message="Updated")

    patch_entity.__annotations__["payload"] = update_schema

    @router.delete(
        "/{entity_id}",
        status_code=status.HTTP_204_NO_CONTENT,
        summary=f"Delete {entity_label}",
        description=f"Delete a {entity_label} record by ID.",
    )
    async def delete_entity(
        entity_id: int,
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        deleted = await repository.delete(entity_id)
        if not deleted:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Resource not found")
        return

    @router.post(
        "/search",
        response_model=APIResponse[SearchResponse[list_schema]],
        summary=f"Search {tags[0]}",
        description=f"Search {tags[0]} records by text query.",
    )
    async def search_entities(
        payload: SearchRequest,
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        results = await repository.search(payload.query)
        response = SearchResponse[list_schema](
            items=[list_schema.model_validate(_orm_to_dict_for_schema(item, list_schema)) for item in results],
            total=len(results),
        )
        return APIResponse(data=response)

    @router.post(
        "/filter",
        response_model=APIResponse[PaginationResponse[list_schema]],
        summary=f"Filter {tags[0]}",
        description=f"Filter {tags[0]} records by a field condition.",
    )
    async def filter_entities(
        payload: FilterRequest,
        pagination: PaginationRequest = Depends(),
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        all_items = await repository.get_all()
        filtered = filter_items(all_items, payload)
        start = (pagination.page - 1) * pagination.page_size
        end = start + pagination.page_size
        response = PaginationResponse[list_schema](
            items=[list_schema.model_validate(_orm_to_dict_for_schema(item, list_schema)) for item in filtered[start:end]],
            total=len(filtered),
            page=pagination.page,
            page_size=pagination.page_size,
            pages=(len(filtered) + pagination.page_size - 1) // pagination.page_size,
        )
        return APIResponse(data=response)

    @router.post(
        "/bulk",
        response_model=APIResponse[list[list_schema]],
        summary=f"Bulk create {tags[0]}",
        description=f"Create multiple {tags[0]} records in bulk.",
    )
    async def bulk_create_entities(
        payload: Any,
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        entities = []
        repo_session = getattr(repository, "session", None)
        for item in payload:
            item_data = item.model_dump()
            filtered = {}
            for key, value in item_data.items():
                if _is_relationship_attribute(model_class, key):
                    fk_id = _resolve_relationship(key, value, repo_session)
                    if fk_id is not None:
                        try:
                            prop = class_mapper(model_class).get_property(key)
                            local_cols = list(prop.local_columns)
                            if local_cols:
                                fk_name = local_cols[0].key
                                filtered[fk_name] = fk_id
                        except Exception:
                            continue
                    continue
                filtered[key] = _normalize_enum_value(key, value)
            entities.append(model_class(**filtered))
        created = await repository.bulk_create(entities)
        return APIResponse(data=[list_schema.model_validate(_orm_to_dict_for_schema(item, list_schema)) for item in created], message="Bulk created")

    bulk_create_entities.__annotations__["payload"] = list[create_schema]

    @router.post(
        "/import",
        response_model=APIResponse[list[list_schema]],
        summary=f"Import {tags[0]} from spreadsheet",
        description=f"Import multiple {tags[0]} records from a CSV or XLSX upload.",
    )
    async def import_entities(
        file: UploadFile = File(...),
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        rows = await _read_import_rows(file)
        if not rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file contains no data rows.")

        entities = []
        repo_session = getattr(repository, "session", None)

        for index, row in enumerate(rows, start=2):
            try:
                payload_obj = create_schema.model_validate(row)
            except ValidationError as exc:
                error_messages = [f"{err['loc'][0]}: {err['msg']}" for err in exc.errors()]
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=f"Invalid data on row {index}: {'; '.join(error_messages)}",
                )

            item_data = payload_obj.model_dump()
            filtered = {}
            for key, value in item_data.items():
                if _is_relationship_attribute(model_class, key):
                    fk_id = _resolve_relationship(key, value, repo_session)
                    if fk_id is not None:
                        try:
                            prop = class_mapper(model_class).get_property(key)
                            local_cols = list(prop.local_columns)
                            if local_cols:
                                fk_name = local_cols[0].key
                                filtered[fk_name] = fk_id
                        except Exception:
                            continue
                    continue
                filtered[key] = _normalize_enum_value(key, value)
            entities.append(model_class(**filtered))

        created = await repository.bulk_create(entities)
        return APIResponse(data=[list_schema.model_validate(_orm_to_dict_for_schema(item, list_schema)) for item in created], message="Imported")

    bulk_create_entities.__annotations__["payload"] = list[create_schema]

    @router.patch(
        "/bulk",
        response_model=APIResponse[list[list_schema]],
        summary=f"Bulk update {tags[0]}",
        description=f"Update multiple {tags[0]} records in bulk.",
    )
    async def bulk_update_entities(
        payload: Any,
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        updated_entities: list[Any] = []
        for item in payload:
            existing = await repository.get_by_id(item.id)
            if existing is None:
                continue
            update_data = item.model_dump(exclude_unset=True)
            update_data.pop("id", None)
            normalized_update = {key: _normalize_enum_value(key, value) for key, value in update_data.items()}
            apply_patch(existing, normalized_update)
            updated_entities.append(existing)
        await repository._commit()
        for entity in updated_entities:
            await repository._refresh(entity)
        return APIResponse(data=[list_schema.model_validate(_orm_to_dict_for_schema(item, list_schema)) for item in updated_entities], message="Bulk updated")

    bulk_update_entities.__annotations__["payload"] = list[bulk_update_schema]

    @router.delete(
        "/bulk",
        response_model=APIResponse[dict[str, int]],
        summary=f"Bulk delete {tags[0]}",
        description=f"Delete multiple {tags[0]} records in bulk.",
    )
    async def bulk_delete_entities(
        payload: BulkOperationRequest,
        repository: BaseRepository[Any] = Depends(repository_dependency),
    ):
        deleted_count = await repository.bulk_delete(payload.ids)
        return APIResponse(data={"deleted": deleted_count}, message="Bulk deleted")

    return router
